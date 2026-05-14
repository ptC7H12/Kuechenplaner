#!/usr/bin/env python3
"""
Excel Recipe Import Script
Importiert Rezepte aus einer Excel-Datei in die Freizeit Rezepturverwaltung DB
"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook

# Add app directory to path
sys.path.insert(0, str(Path(__file__).parent))

from build_logging import setup_build_log
from app.database import SessionLocal, create_tables, SQLALCHEMY_DATABASE_URL
from app import crud, schemas, models

# Spiegele print()-Output nach logs/build_excel_import_*.log
EXCEL_IMPORT_LOG_PATH = setup_build_log("excel_import", Path(__file__).parent / "logs")
print(f"Log: {EXCEL_IMPORT_LOG_PATH}")


def import_recipe_from_sheet(db, sheet):
    """Importiert ein einzelnes Rezept aus einem Excel-Tabellenblatt"""

    # Rezeptname aus A1
    recipe_name = sheet['A1'].value
    if not recipe_name:
        print(f"⚠️  Überspringe Blatt '{sheet.title}' - kein Rezeptname in A1")
        return None

    print(f"\n📖 Importiere Rezept: {recipe_name}")

    # Basisportionen aus A4
    base_servings = sheet['A4'].value
    if not base_servings or not isinstance(base_servings, (int, float)):
        print(f"   ⚠️  Keine gültige Basisportionen in A4, nutze Standard: 30")
        base_servings = 30
    else:
        base_servings = int(base_servings)

    print(f"   Basisportionen: {base_servings}")

    # Zutaten ab Zeile 5 bis Zeile 30
    ingredients = []
    for row in range(5, 31):
        quantity_cell = sheet[f'A{row}'].value
        unit_cell = sheet[f'C{row}'].value
        ingredient_cell = sheet[f'D{row}'].value

        # Abbrechen wenn keine Zutat mehr vorhanden
        if not ingredient_cell:
            break

        # Überspringe Zeilen ohne Menge
        if not quantity_cell:
            continue

        # Parse Menge
        try:
            if isinstance(quantity_cell, str):
                # Ersetze Komma durch Punkt für float parsing
                quantity_cell = quantity_cell.replace(',', '.')
            quantity = float(quantity_cell)
        except (ValueError, TypeError):
            print(f"   ⚠️  Ungültige Menge in Zeile {row}: {quantity_cell}")
            continue

        # Einheit
        unit = str(unit_cell).strip() if unit_cell else "Stück"

        # Zutat Name
        ingredient_name = str(ingredient_cell).strip()

        # Versuche Kategorie zu erraten basierend auf Zutat
        category = guess_ingredient_category(ingredient_name)

        # Hole oder erstelle Zutat in DB
        db_ingredient = crud.get_or_create_ingredient(
            db,
            name=ingredient_name,
            unit=unit,
            category=category
        )

        ingredients.append({
            'ingredient_id': db_ingredient.id,
            'quantity': quantity,
            'unit': unit
        })

        print(f"   ✓ {quantity} {unit} {ingredient_name} ({category})")

    # Anleitung ab Zeile 31
    instructions_lines = []
    for row in range(31, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value:
            instructions_lines.append(str(cell_value).strip())

    instructions = "\n".join(instructions_lines) if instructions_lines else None

    if instructions:
        print(f"   ✓ Anleitung: {len(instructions_lines)} Zeilen")

    # Erstelle Rezept in DB
    try:
        recipe_data = schemas.RecipeCreate(
            name=recipe_name,
            base_servings=base_servings,
            instructions=instructions,
            ingredients=[
                schemas.RecipeIngredientCreate(**ing) for ing in ingredients
            ],
            tag_ids=[],
            allergen_ids=[]
        )

        db_recipe = crud.create_recipe(db, recipe_data)
        print(f"   ✅ Rezept erfolgreich importiert (ID: {db_recipe.id})")
        return db_recipe

    except Exception as e:
        print(f"   ❌ Fehler beim Speichern: {e}")
        return None


def guess_ingredient_category(ingredient_name: str) -> str:
    """Errät die Kategorie einer Zutat basierend auf dem Namen"""

    ingredient_lower = ingredient_name.lower()

    # Gemüse
    if any(veg in ingredient_lower for veg in [
        'kartoffel', 'zwiebel', 'knoblauch', 'tomat', 'gurke', 'paprika',
        'möhre', 'karotte', 'sellerie', 'lauch', 'zucchini', 'aubergine',
        'brokkoli', 'blumenkohl', 'kohl', 'salat', 'spinat', 'erbsen'
    ]):
        return "Gemüse"

    # Obst
    if any(fruit in ingredient_lower for fruit in [
        'apfel', 'birne', 'banane', 'orange', 'zitrone', 'beeren',
        'erdbeere', 'himbeere', 'kirsche', 'pflaume', 'pfirsich'
    ]):
        return "Obst"

    # Fleisch
    if any(meat in ingredient_lower for meat in [
        'fleisch', 'hack', 'rind', 'schwein', 'hähnchen', 'huhn',
        'pute', 'schnitzel', 'würstchen', 'wurst', 'speck', 'schinken'
    ]):
        return "Fleisch"

    # Fisch
    if any(fish in ingredient_lower for fish in [
        'fisch', 'lachs', 'thunfisch', 'forelle', 'garnele', 'krabbe'
    ]):
        return "Fisch"

    # Milchprodukte
    if any(dairy in ingredient_lower for dairy in [
        'milch', 'sahne', 'butter', 'käse', 'quark', 'joghurt',
        'schmand', 'creme', 'frischkäse'
    ]):
        return "Milchprodukte"

    # Eier
    if 'ei' in ingredient_lower or 'eier' in ingredient_lower:
        return "Milchprodukte"

    # Getreide/Teigwaren
    if any(grain in ingredient_lower for grain in [
        'mehl', 'reis', 'nudel', 'pasta', 'brot', 'brötchen',
        'haferflocken', 'müsli', 'couscous', 'bulgur'
    ]):
        return "Getreide"

    # Backwaren/Backzutaten
    if any(baking in ingredient_lower for baking in [
        'zucker', 'backpulver', 'hefe', 'vanille', 'kakao'
    ]):
        return "Backwaren"

    # Öle & Fette
    if any(oil in ingredient_lower for oil in [
        'öl', 'olivenöl', 'sonnenblumenöl', 'margarine', 'fett'
    ]):
        return "Öle & Fette"

    # Gewürze
    if any(spice in ingredient_lower for spice in [
        'salz', 'pfeffer', 'paprika', 'curry', 'zimt', 'muskat',
        'kräuter', 'petersilie', 'basilikum', 'oregano', 'thymian',
        'rosmarin', 'majoran', 'kümmel', 'koriander'
    ]):
        return "Gewürze"

    # Konserven/Haltbar
    if any(canned in ingredient_lower for canned in [
        'dose', 'konserve', 'passiert', 'geschält', 'tomatenmark'
    ]):
        return "Konserven"

    # Sonstiges als Standard
    return "Sonstiges"


def import_recipes_from_excel(excel_path: str):
    """Hauptfunktion zum Importieren aller Rezepte aus einer Excel-Datei"""

    print("=" * 60)
    print("🍳 Excel Rezept Import")
    print("=" * 60)
    print(f"\n💾 Datenbank: {SQLALCHEMY_DATABASE_URL}")

    # Prüfe ob Datei existiert
    if not os.path.exists(excel_path):
        print(f"❌ Datei nicht gefunden: {excel_path}")
        input("\nDrücke Enter zum Beenden...")
        return

    print(f"\n📂 Lade Excel-Datei: {excel_path}")

    try:
        workbook = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ Fehler beim Laden der Excel-Datei: {e}")
        input("\nDrücke Enter zum Beenden...")
        return

    print(f"✓ Gefundene Tabellenblätter: {len(workbook.sheetnames)}")

    # Erstelle DB Session
    db = SessionLocal()

    try:
        # Erstelle Tabellen falls nicht vorhanden
        create_tables()

        imported_count = 0
        skipped_count = 0

        # Durchlaufe alle Tabellenblätter
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            result = import_recipe_from_sheet(db, sheet)

            if result:
                imported_count += 1
            else:
                skipped_count += 1

        print("\n" + "=" * 60)
        print(f"✅ Import abgeschlossen!")
        print(f"   Erfolgreich importiert: {imported_count}")
        print(f"   Übersprungen: {skipped_count}")
        print("=" * 60)
        input("\nDrücke Enter zum Beenden...")

    except Exception as e:
        print(f"\n❌ Fehler beim Import: {e}")
        import traceback
        traceback.print_exc()
        input("\nDrücke Enter zum Beenden...")

    finally:
        db.close()


if __name__ == "__main__":
    try:
        if len(sys.argv) < 2:
            print("Usage: python import_recipes.py <path_to_excel_file>")
            print("\nBeispiel:")
            print("  python import_recipes.py rezepte.xlsx")
            input("\nDrücke Enter zum Beenden...")
            sys.exit(1)

        excel_file = sys.argv[1]
        import_recipes_from_excel(excel_file)
    except Exception as e:
        print(f"\n❌ Unerwarteter Fehler: {e}")
        import traceback
        traceback.print_exc()
        input("\nDrücke Enter zum Beenden...")
        sys.exit(1)