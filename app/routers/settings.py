import contextlib
import json
import os
import shutil
import sys
import tempfile
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, Response
from sqlalchemy.orm import Session

from app import crud, models, schemas
from app.constants import (
    EXCEL_INGREDIENT_ROW_END,
    EXCEL_INGREDIENT_ROW_START,
    EXCEL_INSTRUCTION_ROW_START,
    REPO_URL,
)
from app.database import (
    BACKUP_DIR,
    DATA_DIR,
    _validate_backup_file,
    backup_database,
    get_db,
    reset_database,
    restore_database,
)
from app.dependencies import get_current_camp, get_template_context, templates
from app.logging_config import get_logger
from app.services import unit_converter, update_checker

logger = get_logger("settings")

router = APIRouter()


def _version_candidate_paths() -> list[Path]:
    """Return ordered candidate paths for ``version.txt`` across deployment layouts.

    Layouts considered:
    - Dev / embedded-Python standalone: repo / package root two levels above this file.
    - Nuitka standalone: ``version.txt`` placed next to ``KuechenApp.exe`` inside
      ``_internal/`` (via ``--include-data-file=version.txt=version.txt``).
    - Nuitka standalone with file one level above ``_internal/`` (installer layout
      variant) — checked as defensive fallback.
    """
    candidates: list[Path] = []
    with contextlib.suppress(OSError, IndexError):
        candidates.append(Path(__file__).resolve().parents[2] / "version.txt")
    with contextlib.suppress(OSError):
        exe_dir = Path(sys.executable).resolve().parent
        candidates.append(exe_dir / "version.txt")
        candidates.append(exe_dir.parent / "version.txt")
    return candidates


def _read_version() -> str:
    override = os.environ.get("APP_VERSION")
    if override:
        return override.strip()
    for candidate in _version_candidate_paths():
        try:
            return candidate.read_text(encoding="utf-8").strip()
        except OSError:
            continue
    return "unknown"


APP_VERSION = _read_version()


def safe_json_load(value: str) -> Any:
    """Safely parse JSON or return string value"""
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError, ValueError):
        return value


@router.get("/", response_class=HTMLResponse)
async def settings_page(
    request: Request,
    context: dict = Depends(get_template_context),
    current_camp: models.Camp = Depends(get_current_camp),
    db: Session = Depends(get_db),
):
    """Settings page"""
    all_settings = crud.get_all_settings(db)
    settings_dict = {setting.key: safe_json_load(setting.value) for setting in all_settings}

    tags = crud.get_tags(db)
    allergens = crud.get_allergens(db)
    camps = crud.get_camps(db)
    categories = crud.get_categories(db)
    ingredients = crud.get_ingredients(db, limit=1000)

    context.update(
        {
            "settings": settings_dict,
            "tags": tags,
            "allergens": allergens,
            "camps": camps,
            "categories": categories,
            "ingredients": ingredients,
            "current_camp": current_camp,
            "app_version": APP_VERSION,
            "repo_url": REPO_URL,
            "backup_dir": str(BACKUP_DIR),
            "unit_conversions": unit_converter.load_custom_conversions(db),
        }
    )

    return templates.TemplateResponse("settings/index.html", context)


@router.get("/api/update-check")
async def update_check() -> dict:
    """Return cached GitHub release update info (no remote call within the cache interval)."""
    return update_checker.check_for_update(current_version=APP_VERSION)


@router.get("/database/download")
async def download_database():
    """Create a manual backup of the SQLite database and stream it to the user."""
    backup_path = backup_database(trigger="manual")
    if backup_path is None or not backup_path.exists():
        raise HTTPException(status_code=500, detail="Backup konnte nicht erstellt werden.")
    return FileResponse(
        path=str(backup_path),
        filename=backup_path.name,
        media_type="application/octet-stream",
    )


def _set_flash_toast(response: Response, message: str, toast_type: str = "info") -> None:
    """Set a one-shot toast cookie consumed by base.html after the next page load."""
    payload = urllib.parse.quote(json.dumps({"message": message, "type": toast_type}))
    response.set_cookie(
        key="flash_toast",
        value=payload,
        max_age=30,
        path="/",
        samesite="lax",
    )


def _toast_trigger_header(message: str, toast_type: str = "error") -> dict[str, str]:
    """Return an HX-Trigger header dict so HTMX can show an inline toast."""
    return {"HX-Trigger": json.dumps({"flash-toast": {"message": message, "type": toast_type}})}


@router.post("/database/restore")
async def restore_database_endpoint(
    backup_file: UploadFile = File(...),
):
    """Replace the live database with an uploaded `.db` backup.

    Workflow: save upload to a temp file in `DATA_DIR`, validate (SQLite header
    + alembic_version with known revision), then swap and re-run migrations.
    On success the current_camp_id cookie is cleared (it may point at an
    obsolete camp) and the client is forced to reload.
    """
    if not backup_file.filename or not backup_file.filename.lower().endswith(".db"):
        return Response(
            status_code=400,
            headers=_toast_trigger_header("Bitte eine .db-Backup-Datei auswählen.", "error"),
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp_path = DATA_DIR / f"app_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.tmp"
    try:
        with tmp_path.open("wb") as out:
            shutil.copyfileobj(backup_file.file, out)
    finally:
        await backup_file.close()

    try:
        _validate_backup_file(tmp_path)
    except ValueError as exc:
        tmp_path.unlink(missing_ok=True)
        logger.warning("Restore rejected: %s", exc)
        return Response(status_code=400, headers=_toast_trigger_header(str(exc), "error"))

    try:
        restore_database(tmp_path)
    except Exception as exc:
        logger.exception("Restore failed")
        tmp_path.unlink(missing_ok=True)
        return Response(
            status_code=500,
            headers=_toast_trigger_header(f"Wiederherstellung fehlgeschlagen: {exc}", "error"),
        )

    response = Response(headers={"HX-Refresh": "true"})
    response.delete_cookie("current_camp_id", path="/")
    _set_flash_toast(response, "Backup wiederhergestellt.", "success")
    return response


@router.post("/database/reset")
async def reset_database_endpoint():
    """Delete the live database, re-run migrations, and force a reload."""
    try:
        reset_database()
    except Exception as exc:
        logger.exception("Reset failed")
        return Response(
            status_code=500,
            headers=_toast_trigger_header(f"Zurücksetzen fehlgeschlagen: {exc}", "error"),
        )

    response = Response(headers={"HX-Refresh": "true"})
    response.delete_cookie("current_camp_id", path="/")
    _set_flash_toast(response, "Datenbank zurückgesetzt.", "success")
    return response


@router.get("/api/settings")
async def get_all_settings(db: Session = Depends(get_db)):
    """Get all settings (API endpoint)"""
    settings = crud.get_all_settings(db)
    return {setting.key: crud.get_setting_value(db, setting.key) for setting in settings}


@router.get("/api/settings/{key}")
async def get_setting(key: str, db: Session = Depends(get_db)):
    """Get a specific setting (API endpoint)"""
    value = crud.get_setting_value(db, key)
    if value is None:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")

    return {"key": key, "value": value}


@router.post("/api/settings")
async def update_setting(key: str, value: Any, db: Session = Depends(get_db)):
    """Update or create a setting (API endpoint)"""
    crud.set_setting_value(db, key, value)
    return {"success": True, "key": key, "value": value}


@router.put("/api/settings/{key}")
async def update_specific_setting(key: str, value: dict[str, Any], db: Session = Depends(get_db)):
    """Update a specific setting (API endpoint)"""
    crud.set_setting_value(db, key, value.get("value"))
    return {"success": True, "key": key, "value": value.get("value")}


@router.delete("/api/settings/{key}")
async def delete_setting(key: str, db: Session = Depends(get_db)):
    """Delete a setting (API endpoint)"""
    setting = crud.delete_setting(db, key)
    if not setting:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")

    return {"success": True, "message": f"Setting '{key}' deleted"}


# Unit conversion settings
@router.get("/api/settings/units/conversions")
async def get_unit_conversions(db: Session = Depends(get_db)):
    """Get unit conversion settings"""
    return crud.get_setting_value(db, "unit_conversions", default={})


@router.post("/api/settings/units/conversions")
async def update_unit_conversions(conversions: dict[str, Any], db: Session = Depends(get_db)):
    """Update unit conversion settings"""
    crud.set_setting_value(db, "unit_conversions", conversions)
    return {"success": True, "conversions": conversions}


def _render_conversions_list(request: Request, db: Session) -> HTMLResponse:
    """Render the custom-conversions list partial as an HTMX fragment."""
    return templates.TemplateResponse(
        "settings/_conversions_list.html",
        {"request": request, "unit_conversions": unit_converter.load_custom_conversions(db)},
    )


@router.post("/api/settings/units/conversions/add")
async def add_unit_conversion(request: Request, db: Session = Depends(get_db)):
    """Add a single custom unit conversion and return the updated list partial.

    Threshold is always 1, so the conversion applies to any quantity. The
    underlying service overwrites an existing rule for the same `from_unit`.
    """
    form_data = await request.form()
    from_unit = (form_data.get("from_unit") or "").strip()
    to_unit = (form_data.get("to_unit") or "").strip()
    factor_raw = (form_data.get("factor") or "").strip()

    if not from_unit or not to_unit or not factor_raw:
        return HTMLResponse(
            status_code=400,
            headers=_toast_trigger_header("Bitte alle Felder ausfüllen.", "error"),
            content="",
        )

    try:
        factor = float(factor_raw.replace(",", "."))
    except ValueError:
        return HTMLResponse(
            status_code=400,
            headers=_toast_trigger_header("Faktor muss eine Zahl sein.", "error"),
            content="",
        )

    if factor <= 0:
        return HTMLResponse(
            status_code=400,
            headers=_toast_trigger_header("Faktor muss eine positive Zahl sein.", "error"),
            content="",
        )

    unit_converter.add_custom_conversion(db, from_unit, to_unit, threshold=1, factor=factor)

    response = _render_conversions_list(request, db)
    response.headers["HX-Trigger"] = json.dumps(
        {"flash-toast": {"message": "Konvertierung gespeichert.", "type": "success"}}
    )
    return response


@router.delete("/api/settings/units/conversions/{from_unit}")
async def delete_unit_conversion(from_unit: str, request: Request, db: Session = Depends(get_db)):
    """Remove a single custom unit conversion and return the updated list partial."""
    unit_converter.remove_custom_conversion(db, from_unit)
    return _render_conversions_list(request, db)


# Tag management endpoints
@router.post("/api/tags")
async def create_tag(request: Request, db: Session = Depends(get_db)):
    """Create a new tag"""
    form_data = await request.form()
    name = form_data.get("name")
    icon = form_data.get("icon", "🏷️")
    color = form_data.get("color", "#3B82F6")

    if not name:
        raise HTTPException(status_code=400, detail="Tag name is required")

    existing_tag = crud.get_tag_by_name(db, name)
    if existing_tag:
        raise HTTPException(status_code=400, detail="Tag already exists")

    tag_data = schemas.TagCreate(name=name, icon=icon, color=color)
    tag = crud.create_tag(db, tag_data)

    html = f"""
    <div class="flex items-center justify-between p-4 border-2 border-gray-200 rounded-xl bg-white hover:shadow-lg transition-all" id="tag-{tag.id}">
        <div class="flex items-center space-x-3">
            <span class="text-2xl">{tag.icon or "🏷️"}</span>
            <span class="font-semibold text-gray-900">{tag.name}</span>
            <div class="w-5 h-5 rounded-full border-2 border-gray-200 shadow-sm" style="background-color: {tag.color}"></div>
        </div>
        <button hx-delete="/settings/api/tags/{tag.id}"
                hx-confirm="Tag '{tag.name}' wirklich löschen?"
                hx-target="#tag-{tag.id}"
                hx-swap="outerHTML swap:0.3s"
                class="text-red-600 hover:text-red-800 hover:bg-red-50 p-2 rounded-lg transition-all">
            <svg class="w-5 h-5" fill="currentColor" viewBox="0 0 20 20">
                <path fill-rule="evenodd" d="M9 2a1 1 0 00-.894.553L7.382 4H4a1 1 0 000 2v10a2 2 0 002 2h8a2 2 0 002-2V6a1 1 0 100-2h-3.382l-.724-1.447A1 1 0 0011 2H9zM7 8a1 1 0 012 0v6a1 1 0 11-2 0V8zm5-1a1 1 0 00-1 1v6a1 1 0 102 0V8a1 1 0 00-1-1z" clip-rule="evenodd"></path>
            </svg>
        </button>
    </div>
    """

    return HTMLResponse(content=html, status_code=201)


@router.delete("/api/tags/{tag_id}")
async def delete_tag(tag_id: int, db: Session = Depends(get_db)):
    """Delete a tag"""
    tag = crud.delete_tag(db, tag_id)
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")

    return HTMLResponse(content="", status_code=200)


# Ingredient category management endpoints
@router.post("/api/categories")
async def create_category(request: Request, db: Session = Depends(get_db)):
    """Create a new ingredient category and return its card fragment."""
    form_data = await request.form()
    name = (form_data.get("name") or "").strip()
    color = (form_data.get("color") or "#6B7280").strip()

    if not name:
        raise HTTPException(status_code=422, detail="Kategoriename ist erforderlich")

    if crud.get_category_by_name(db, name):
        raise HTTPException(status_code=400, detail="Kategorie existiert bereits")

    category = crud.create_category(db, schemas.CategoryCreate(name=name, color=color))
    return templates.TemplateResponse(
        "settings/_category_card.html",
        {"request": request, "category": category},
        status_code=201,
    )


@router.delete("/api/categories/{category_id}")
async def delete_category(category_id: int, db: Session = Depends(get_db)):
    """Delete a category; referencing ingredients fall back to no category."""
    category = crud.delete_category(db, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Kategorie nicht gefunden")

    return HTMLResponse(content="", status_code=200)


# Allergen management endpoints
@router.post("/api/allergens")
async def create_allergen(request: Request, db: Session = Depends(get_db)):
    """Create a new allergen and return its card fragment."""
    form_data = await request.form()
    name = (form_data.get("name") or "").strip()
    icon = (form_data.get("icon") or "").strip() or None

    if not name:
        raise HTTPException(status_code=422, detail="Allergen-Name ist erforderlich")

    if crud.get_allergen_by_name(db, name):
        raise HTTPException(status_code=400, detail="Allergen existiert bereits")

    allergen = crud.create_allergen(db, schemas.AllergenCreate(name=name, icon=icon))
    return templates.TemplateResponse(
        "settings/_allergen_card.html",
        {"request": request, "allergen": allergen},
        status_code=201,
    )


@router.delete("/api/allergens/{allergen_id}")
async def delete_allergen(allergen_id: int, db: Session = Depends(get_db)):
    """Delete an allergen; existing recipe links are removed via the M:N table."""
    allergen = crud.delete_allergen(db, allergen_id)
    if not allergen:
        raise HTTPException(status_code=404, detail="Allergen nicht gefunden")

    return HTMLResponse(content="", status_code=200)


# Ingredient management endpoints
@router.get("/ingredients", response_class=HTMLResponse)
async def list_ingredients(request: Request, search: str = "", db: Session = Depends(get_db)):
    """Render the ingredient management table (HTMX search fragment)."""
    ingredients = crud.get_ingredients(db, search=search or None, limit=1000)
    categories = crud.get_categories(db)
    return templates.TemplateResponse(
        "settings/_ingredients_table.html",
        {"request": request, "ingredients": ingredients, "categories": categories},
    )


@router.patch("/ingredients/{ingredient_id}")
async def update_ingredient(ingredient_id: int, request: Request, db: Session = Depends(get_db)):
    """Update an ingredient's category and/or unit; return the refreshed row."""
    form_data = await request.form()
    raw_category_id = (form_data.get("category_id") or "").strip()
    unit = (form_data.get("unit") or "").strip()

    update_kwargs: dict[str, Any] = {"category_id": int(raw_category_id) if raw_category_id else None}
    if unit:
        update_kwargs["unit"] = unit

    ingredient = crud.update_ingredient(db, ingredient_id, schemas.IngredientUpdate(**update_kwargs))
    if not ingredient:
        raise HTTPException(status_code=404, detail="Zutat nicht gefunden")

    categories = crud.get_categories(db)
    return templates.TemplateResponse(
        "settings/_ingredient_row.html",
        {"request": request, "ingredient": ingredient, "categories": categories},
    )


# Excel Recipe Import
def _guess_ingredient_category(ingredient_name: str) -> str:
    """Guess ingredient category based on name"""
    ingredient_lower = ingredient_name.lower()

    categories = {
        "Gemüse": [
            "kartoffel",
            "zwiebel",
            "knoblauch",
            "tomat",
            "gurke",
            "paprika",
            "möhre",
            "karotte",
            "sellerie",
            "lauch",
            "zucchini",
            "aubergine",
            "brokkoli",
            "blumenkohl",
            "kohl",
            "salat",
            "spinat",
            "erbsen",
        ],
        "Obst": [
            "apfel",
            "birne",
            "banane",
            "orange",
            "zitrone",
            "beeren",
            "erdbeere",
            "himbeere",
            "kirsche",
            "pflaume",
            "pfirsich",
        ],
        "Fleisch": [
            "fleisch",
            "hack",
            "rind",
            "schwein",
            "hähnchen",
            "huhn",
            "pute",
            "schnitzel",
            "würstchen",
            "wurst",
            "speck",
            "schinken",
        ],
        "Fisch": ["fisch", "lachs", "thunfisch", "forelle", "garnele", "krabbe"],
        "Milchprodukte": ["milch", "sahne", "butter", "käse", "quark", "joghurt", "schmand", "creme", "frischkäse"],
        "Getreide": [
            "mehl",
            "reis",
            "nudel",
            "pasta",
            "brot",
            "brötchen",
            "haferflocken",
            "müsli",
            "couscous",
            "bulgur",
        ],
        "Backwaren": ["zucker", "backpulver", "hefe", "vanille", "kakao"],
        "Öle & Fette": ["öl", "olivenöl", "sonnenblumenöl", "margarine", "fett"],
        "Gewürze": [
            "salz",
            "pfeffer",
            "paprika",
            "curry",
            "zimt",
            "muskat",
            "kräuter",
            "petersilie",
            "basilikum",
            "oregano",
            "thymian",
            "rosmarin",
            "majoran",
            "kümmel",
            "koriander",
        ],
        "Konserven": ["dose", "konserve", "passiert", "geschält", "tomatenmark"],
    }

    for category, keywords in categories.items():
        if any(kw in ingredient_lower for kw in keywords):
            return category

    if "ei" in ingredient_lower or "eier" in ingredient_lower:
        return "Milchprodukte"

    return "Sonstiges"


def _import_recipe_from_sheet(db: Session, sheet):
    """Import a single recipe from an Excel sheet"""
    recipe_name = sheet["A1"].value
    if not recipe_name:
        return None, f"Blatt '{sheet.title}': Kein Rezeptname in A1"

    recipe_name = str(recipe_name).strip()

    # Check for duplicate
    existing = db.query(models.Recipe).filter(models.Recipe.name == recipe_name).first()
    if existing:
        return None, f"'{recipe_name}': Existiert bereits"

    base_servings = sheet["A4"].value
    base_servings = 30 if not base_servings or not isinstance(base_servings, (int, float)) else int(base_servings)

    ingredients = []
    for row in range(EXCEL_INGREDIENT_ROW_START, EXCEL_INGREDIENT_ROW_END):
        quantity_cell = sheet[f"A{row}"].value
        unit_cell = sheet[f"C{row}"].value
        ingredient_cell = sheet[f"D{row}"].value

        if not ingredient_cell:
            break
        if not quantity_cell:
            continue

        try:
            if isinstance(quantity_cell, str):
                quantity_cell = quantity_cell.replace(",", ".")
            quantity = float(quantity_cell)
        except (ValueError, TypeError):
            continue

        unit = str(unit_cell).strip() if unit_cell else "Stück"
        ingredient_name = str(ingredient_cell).strip()
        category = _guess_ingredient_category(ingredient_name)

        db_ingredient = crud.get_or_create_ingredient(db, name=ingredient_name, unit=unit, category=category)
        ingredients.append({"ingredient_id": db_ingredient.id, "quantity": quantity, "unit": unit})

    instructions_lines = []
    for row in range(EXCEL_INSTRUCTION_ROW_START, sheet.max_row + 1):
        cell_value = sheet[f"A{row}"].value
        if cell_value:
            instructions_lines.append(str(cell_value).strip())

    instructions = "\n".join(instructions_lines) if instructions_lines else None

    recipe_data = schemas.RecipeCreate(
        name=recipe_name,
        base_servings=base_servings,
        instructions=instructions,
        ingredients=[schemas.RecipeIngredientCreate(**ing) for ing in ingredients],
        tag_ids=[],
        allergen_ids=[],
    )

    db_recipe = crud.create_recipe(db, recipe_data)
    return db_recipe, None


@router.post("/api/import-recipes")
async def import_recipes_from_excel(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import recipes from an Excel file"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return HTMLResponse(
            content='<div class="alert alert-error">Bitte eine Excel-Datei (.xlsx) hochladen.</div>', status_code=400
        )

    try:
        from openpyxl import load_workbook

        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            workbook = load_workbook(tmp_path, data_only=True)
        finally:
            os.unlink(tmp_path)

        imported = []
        skipped = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            try:
                recipe, error = _import_recipe_from_sheet(db, sheet)
                if recipe:
                    imported.append(recipe.name)
                elif error:
                    skipped.append(error)
            except Exception as e:
                logger.exception("Sheet import failed: %s", sheet_name)
                skipped.append(f"'{sheet_name}': {str(e)}")

        logger.info(f"Excel import: {len(imported)} imported, {len(skipped)} skipped")

        # Build result HTML
        result_parts = []
        if imported:
            recipes_list = "".join(f"<li>{name}</li>" for name in imported)
            result_parts.append(
                f'<div class="alert alert-success mb-3">'
                f"<strong>{len(imported)} Rezept(e) erfolgreich importiert:</strong>"
                f'<ul class="list-disc ml-6 mt-2">{recipes_list}</ul></div>'
            )
        if skipped:
            skip_list = "".join(f"<li>{msg}</li>" for msg in skipped)
            result_parts.append(
                f'<div class="alert alert-warning">'
                f"<strong>{len(skipped)} übersprungen:</strong>"
                f'<ul class="list-disc ml-6 mt-2">{skip_list}</ul></div>'
            )
        if not imported and not skipped:
            result_parts.append('<div class="alert alert-warning">Keine Rezepte in der Datei gefunden.</div>')

        return HTMLResponse(content="".join(result_parts))

    except Exception as e:
        logger.error(f"Excel import error: {e}", exc_info=True)
        return HTMLResponse(
            content=f'<div class="alert alert-error">Fehler beim Import: {str(e)}</div>', status_code=500
        )
