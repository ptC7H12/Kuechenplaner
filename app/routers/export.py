from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from datetime import datetime, timezone, timedelta
from io import BytesIO
import os
from pathlib import Path
import re
import subprocess
import sys
from collections import defaultdict

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app import crud, models
from app.services.calculation import calculate_shopping_list
from app.services.unit_converter import format_quantity_with_conversion
from app.logging_config import get_logger

logger = get_logger("export")

router = APIRouter()


SUB_CATEGORY_COLORS = {
    "Vorspeise": colors.HexColor("#10b981"),
    "Hauptgang": colors.HexColor("#6366f1"),
    "Beilage": colors.HexColor("#d97706"),
    "Salat": colors.HexColor("#84cc16"),
    "Nachtisch": colors.HexColor("#f97316"),
}
DEFAULT_SUB_CATEGORY_COLOR = colors.HexColor("#6b7280")


# --- Helper functions ---

def sanitize_filename(name: str) -> str:
    """Sanitize a string for safe use as a filename"""
    return re.sub(r'[^\w\-.]', '_', name)


def get_german_weekday(date):
    """Get German weekday name for a date"""
    weekdays = {
        0: "Montag", 1: "Dienstag", 2: "Mittwoch", 3: "Donnerstag",
        4: "Freitag", 5: "Samstag", 6: "Sonntag"
    }
    return weekdays.get(date.weekday(), date.strftime('%A'))


def get_downloads_folder():
    """Get or create downloads folder for exported files"""
    downloads_path = Path.home() / 'Downloads' / 'Kuechenplaner'
    downloads_path.mkdir(parents=True, exist_ok=True)
    return downloads_path


def open_file(filepath):
    """Open file with default system viewer"""
    try:
        if sys.platform == 'win32':
            os.startfile(filepath)
        elif sys.platform == 'darwin':
            subprocess.run(['open', filepath])
        else:
            subprocess.run(['xdg-open', filepath])
        return True
    except Exception as e:
        logger.error(f"Error opening file: {e}", exc_info=True)
        return False


def build_and_serve_pdf(buffer: BytesIO, filename: str) -> FileResponse:
    """Save PDF to downloads folder, open it, and return FileResponse"""
    buffer.seek(0)
    downloads_folder = get_downloads_folder()
    filepath = downloads_folder / filename

    with open(filepath, 'wb') as f:
        f.write(buffer.getvalue())

    open_file(str(filepath))

    return FileResponse(
        path=str(filepath),
        media_type="application/pdf",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def get_pdf_styles():
    """Get common PDF styles used across all exports"""
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ExportTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=30
    )

    heading_style = ParagraphStyle(
        'ExportHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12
    )

    section_heading_style = ParagraphStyle(
        'ExportSectionHeading',
        parent=styles['Heading3'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )

    return styles, title_style, heading_style, section_heading_style


def get_table_style():
    """Get common table style for ingredient/data tables"""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ])


def make_timestamp():
    """Generate a timestamp string for filenames"""
    return datetime.now(timezone.utc).strftime('%Y%m%d')


# --- Export endpoints ---

@router.get("/shopping-list/pdf/{camp_id}")
async def export_shopping_list_pdf(
    camp_id: int,
    db: Session = Depends(get_db)
):
    """Export shopping list as PDF (compact layout with notes column)"""
    camp = crud.get_camp(db, camp_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Camp not found")

    shopping_data = calculate_shopping_list(db, camp_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = []
    styles, _, _, _ = get_pdf_styles()

    compact_title = ParagraphStyle(
        'ShoppingListTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=8,
    )
    compact_heading = ParagraphStyle(
        'ShoppingListHeading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        'ShoppingListCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
    )

    # Title
    elements.append(Paragraph(f"Einkaufsliste: {camp.name}", compact_title))

    # Camp info
    info_text = (
        f"Zeitraum: {camp.start_date.strftime('%d.%m.%Y')} - {camp.end_date.strftime('%d.%m.%Y')} | "
        f"Teilnehmer: {camp.participant_count} | "
        f"Rezepte: {shopping_data['total_recipes']} | "
        f"Zutaten: {shopping_data['total_items']}"
    )
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 8))

    compact_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ])

    # Shopping list by category
    for category, items in shopping_data['categories'].items():
        elements.append(Paragraph(category, compact_heading))

        table_data = [["Zutat", "Menge", "Einheit", "Bemerkung"]]
        for item in items:
            note_parts = []
            if item.get('global_note'):
                note_parts.append(str(item['global_note']))
            if item.get('note'):
                note_parts.append(str(item['note']))
            note_text = "<br/>".join(note_parts) if note_parts else ""

            table_data.append([
                Paragraph(item['ingredient'].name, cell_style),
                f"{item['quantity']:.1f}",
                item['unit'],
                Paragraph(note_text, cell_style),
            ])

        table = Table(table_data, colWidths=[6.5*cm, 2.2*cm, 2.0*cm, 7.3*cm])
        table.setStyle(compact_table_style)
        elements.append(table)
        elements.append(Spacer(1, 10))

    doc.build(elements)

    filename = f"einkaufsliste_{sanitize_filename(camp.name)}_{make_timestamp()}.pdf"
    return build_and_serve_pdf(buffer, filename)


@router.get("/shopping-list/excel/{camp_id}")
async def export_shopping_list_excel(
    camp_id: int,
    db: Session = Depends(get_db)
):
    """Export shopping list as Excel"""
    camp = crud.get_camp(db, camp_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Camp not found")

    shopping_data = calculate_shopping_list(db, camp_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Einkaufsliste"

    ws['A1'] = f"Einkaufsliste: {camp.name}"
    ws['A1'].font = Font(size=16, bold=True)

    ws['A3'] = "Zeitraum:"
    ws['B3'] = f"{camp.start_date.strftime('%d.%m.%Y')} - {camp.end_date.strftime('%d.%m.%Y')}"
    ws['A4'] = "Teilnehmer:"
    ws['B4'] = camp.participant_count
    ws['A5'] = "Rezepte:"
    ws['B5'] = shopping_data['total_recipes']

    row = 7
    for col_letter, header in [('A', 'Kategorie'), ('B', 'Zutat'), ('C', 'Menge'), ('D', 'Einheit'), ('E', '?')]:
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        cell.alignment = Alignment(horizontal='left')

    row += 1

    for category, items in shopping_data['categories'].items():
        for item in items:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = item['ingredient'].name
            ws[f'C{row}'] = round(item['quantity'], 1)
            ws[f'D{row}'] = item['unit']
            ws[f'E{row}'] = ""
            row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 5

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"einkaufsliste_{sanitize_filename(camp.name)}_{make_timestamp()}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/meal-plan/pdf/{camp_id}")
async def export_meal_plan_pdf(
    camp_id: int,
    db: Session = Depends(get_db)
):
    """Export meal plan as PDF in landscape table format"""
    camp = crud.get_camp(db, camp_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Camp not found")

    meal_plans = crud.get_meal_plans_for_camp(db, camp_id)

    meal_grid = defaultdict(lambda: {
        models.MealType.BREAKFAST: [],
        models.MealType.LUNCH: [],
        models.MealType.DINNER: []
    })

    for meal_plan in meal_plans:
        date_key = meal_plan.meal_date.date()
        meal_grid[date_key][meal_plan.meal_type].append(meal_plan)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'MealPlanTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    # Generate all days
    all_days = []
    current_date = camp.start_date
    while current_date <= camp.end_date:
        all_days.append(current_date)
        current_date += timedelta(days=1)

    # Split days into pages (10 days per page)
    pages = [all_days[i:i+10] for i in range(0, len(all_days), 10)]

    meal_types = [
        (models.MealType.BREAKFAST, "Frühstück"),
        (models.MealType.LUNCH, "Mittagessen"),
        (models.MealType.DINNER, "Abendessen")
    ]

    for page_idx, page_days in enumerate(pages):
        if page_idx > 0:
            elements.append(PageBreak())

        elements.append(Paragraph(f"Speiseplan: {camp.name}", title_style))

        info_text = f"Zeitraum: {camp.start_date.strftime('%d.%m.%Y')} - {camp.end_date.strftime('%d.%m.%Y')} | Teilnehmer: {camp.participant_count}"
        if len(pages) > 1:
            info_text += f" | Seite {page_idx + 1} von {len(pages)}"

        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 15))

        table_data = [["Datum"] + [meal_name for _, meal_name in meal_types]]

        for day in page_days:
            date_key = day.date()
            weekday = get_german_weekday(day)
            row = [f"{weekday}\n{day.strftime('%d.%m.%Y')}"]

            for meal_type, _ in meal_types:
                recipes = meal_grid[date_key][meal_type]
                if recipes:
                    lines = []
                    for mp in recipes:
                        name = mp.recipe.name if mp.recipe else "Kein Essen"
                        prefix = f"{mp.sub_category}: " if mp.sub_category else ""
                        suffix = f" ({mp.custom_servings} Pers.)" if mp.custom_servings else ""
                        lines.append(f"{prefix}{name}{suffix}")
                    row.append("\n".join(lines))
                else:
                    row.append("-")

            table_data.append(row)

        date_col_width = 3.5*cm
        meal_col_width = (landscape(A4)[0] - 3*cm - date_col_width) / len(meal_types)
        col_widths = [date_col_width] + [meal_col_width] * len(meal_types)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
            ('FONTSIZE', (1, 1), (-1, -1), 7),
            ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (1, 1), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#4F46E5')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))

        elements.append(table)

    doc.build(elements)

    filename = f"speiseplan_{sanitize_filename(camp.name)}_{make_timestamp()}.pdf"
    return build_and_serve_pdf(buffer, filename)


@router.get("/recipe-book/pdf/{camp_id}")
async def export_recipe_book_pdf(
    camp_id: int,
    db: Session = Depends(get_db)
):
    """Export recipe book with all recipes from meal plan.

    Scales each recipe to the largest custom_servings value seen across
    meal_plans referencing it (falls back to camp.participant_count).
    """
    camp = crud.get_camp(db, camp_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Camp not found")

    meal_plans = crud.get_meal_plans_for_camp(db, camp_id)

    # Pick the largest effective servings per recipe so that a single
    # printed copy covers all planned occurrences.
    recipe_servings: dict[int, int] = {}
    for mp in meal_plans:
        if mp.recipe_id is None:
            continue
        effective = mp.custom_servings or camp.participant_count
        prev = recipe_servings.get(mp.recipe_id, 0)
        if effective > prev:
            recipe_servings[mp.recipe_id] = effective

    recipes = [r for r in (crud.get_recipe(db, rid) for rid in recipe_servings) if r is not None]

    if not recipes:
        raise HTTPException(status_code=404, detail="No recipes found in meal plan")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles, title_style, _, section_heading_style = get_pdf_styles()

    elements.append(Paragraph(f"Rezeptbuch: {camp.name}", title_style))
    elements.append(Spacer(1, 20))

    for i, recipe in enumerate(recipes):
        if i > 0:
            elements.append(PageBreak())

        target_servings = recipe_servings[recipe.id]
        scaling_factor = target_servings / recipe.base_servings

        elements.append(Paragraph(recipe.name, styles['Heading1']))

        if recipe.description:
            elements.append(Paragraph(recipe.description, styles['Normal']))
            elements.append(Spacer(1, 10))

        info_text = f"Portionen: {target_servings} (Originalrezept: {recipe.base_servings}) | "
        if recipe.preparation_time:
            info_text += f"Vorbereitung: {recipe.preparation_time} min | "
        if recipe.cooking_time:
            info_text += f"Kochzeit: {recipe.cooking_time} min"

        elements.append(Paragraph(f"<i>{info_text}</i>", styles['Normal']))
        elements.append(Spacer(1, 15))

        if recipe.allergens:
            allergen_names = ", ".join([a.name for a in recipe.allergens])
            elements.append(Paragraph(f"<b>Allergene:</b> {allergen_names}", styles['Normal']))
            elements.append(Spacer(1, 10))

        elements.append(Paragraph("<b>Zutaten:</b>", styles['Heading3']))

        ingredient_data = [["Zutat", "Menge"]]
        for ri in recipe.ingredients:
            ingredient_data.append([
                ri.ingredient.name,
                format_quantity_with_conversion(ri.quantity * scaling_factor, ri.unit),
            ])

        ing_table = Table(ingredient_data, colWidths=[10*cm, 6*cm])
        ing_table.setStyle(get_table_style())
        elements.append(ing_table)
        elements.append(Spacer(1, 15))

        if recipe.instructions:
            elements.append(Paragraph("<b>Zubereitung:</b>", styles['Heading3']))
            elements.append(Paragraph(recipe.instructions, styles['Normal']))

    doc.build(elements)

    filename = f"rezeptbuch_{sanitize_filename(camp.name)}_{make_timestamp()}.pdf"
    return build_and_serve_pdf(buffer, filename)


@router.get("/daily-lists/pdf/{camp_id}")
async def export_daily_lists_pdf(
    camp_id: int,
    db: Session = Depends(get_db)
):
    """Export a "Tageslisten" PDF: one page per camp day, grouped by meal type
    (with sub-categories for dinner), each recipe printed with its ingredients
    and preparation steps in a 3-column table.
    """
    from app.constants import MEAL_SUB_CATEGORIES

    camp = crud.get_camp(db, camp_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Camp not found")

    meal_plans = crud.get_meal_plans_for_camp(db, camp_id)
    if not meal_plans:
        raise HTTPException(status_code=404, detail="No meal plans found")

    # Group by date → meal_type, preserving sub-category ordering for dinner.
    sub_cat_order = {name: idx for idx, name in enumerate(MEAL_SUB_CATEGORIES)}

    def sort_key(mp):
        sub_idx = sub_cat_order.get(mp.sub_category, len(MEAL_SUB_CATEGORIES))
        return (sub_idx, mp.position or 0)

    by_date: dict = defaultdict(lambda: {
        models.MealType.BREAKFAST: [],
        models.MealType.LUNCH: [],
        models.MealType.DINNER: [],
    })
    for mp in meal_plans:
        if not mp.recipe_id or not mp.recipe:
            continue
        by_date[mp.meal_date.date()][mp.meal_type].append(mp)

    for date_key in by_date:
        for mt in by_date[date_key]:
            by_date[date_key][mt].sort(key=sort_key)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = []
    styles, _, _, _ = get_pdf_styles()

    day_title_style = ParagraphStyle(
        'DayTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=10,
        fontName='Helvetica-Bold',
    )
    meal_heading_style = ParagraphStyle(
        'MealHeading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
    )
    cell_style = ParagraphStyle(
        'DailyCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
    )

    meal_type_labels = {
        models.MealType.BREAKFAST: "Frühstück",
        models.MealType.LUNCH: "Mittagessen",
        models.MealType.DINNER: "Abendessen",
    }

    def build_recipe_block(meal_plan, total_width=18.0*cm):
        recipe = meal_plan.recipe
        effective_servings = meal_plan.custom_servings or camp.participant_count
        factor = effective_servings / recipe.base_servings if recipe.base_servings else 1.0

        ingredients = list(recipe.ingredients)
        instructions_text = (recipe.instructions or "").strip()
        has_instructions = bool(instructions_text)
        rows = max(len(ingredients), 1)

        header = ["Menge", "Zutat"]
        if has_instructions:
            header.append("Zubereitung")
        table_data = [header]

        instructions_paragraph = (
            Paragraph(instructions_text.replace("\n", "<br/>"), cell_style)
            if has_instructions else None
        )

        for i in range(rows):
            if i < len(ingredients):
                ri = ingredients[i]
                qty_cell = format_quantity_with_conversion(ri.quantity * factor, ri.unit)
                name_cell = ri.ingredient.name
            else:
                qty_cell = ""
                name_cell = ""

            if has_instructions:
                # Single Paragraph in the first row, empty cells underneath; the
                # SPAN below makes it stretch over all ingredient rows.
                step_cell = instructions_paragraph if i == 0 else ""
                table_data.append([qty_cell, Paragraph(name_cell, cell_style), step_cell])
            else:
                table_data.append([qty_cell, Paragraph(name_cell, cell_style)])

        if has_instructions:
            col_widths = [
                total_width * (2.5 / 18.0),
                total_width * (5.0 / 18.0),
                total_width * (10.5 / 18.0),
            ]
        else:
            col_widths = [
                total_width * (3.0 / 17.5),
                total_width * (14.5 / 17.5),
            ]

        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 1), (0, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]
        if has_instructions and rows > 1:
            style_cmds.append(('SPAN', (2, 1), (2, rows)))

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle(style_cmds))
        return table

    def build_sub_category_block(meal_plan):
        """Indented block with a colored left bar and sub-category heading."""
        color = SUB_CATEGORY_COLORS.get(meal_plan.sub_category, DEFAULT_SUB_CATEGORY_COLOR)
        sub_label = meal_plan.sub_category or "Sonstiges"
        sub_heading_style = ParagraphStyle(
            'SubCategoryHeading',
            parent=styles['Normal'],
            fontSize=11,
            textColor=color,
            fontName='Helvetica-Bold',
            spaceAfter=2,
        )
        heading = Paragraph(f"{sub_label}: {meal_plan.recipe.name}", sub_heading_style)
        inner_table = build_recipe_block(meal_plan, total_width=17.0*cm)

        wrapper = Table(
            [[None, heading], [None, inner_table]],
            colWidths=[1.0*cm, 17.0*cm],
        )
        wrapper.setStyle(TableStyle([
            ('LINEBEFORE', (1, 0), (1, -1), 3, color),
            ('LEFTPADDING', (1, 0), (1, -1), 8),
            ('RIGHTPADDING', (1, 0), (1, -1), 0),
            ('LEFTPADDING', (0, 0), (0, -1), 0),
            ('RIGHTPADDING', (0, 0), (0, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        return wrapper

    sorted_dates = sorted(by_date.keys())
    for idx, day in enumerate(sorted_dates):
        if idx > 0:
            elements.append(PageBreak())

        weekday = get_german_weekday(day)
        elements.append(Paragraph(f"{weekday}, {day.strftime('%d.%m.%Y')}", day_title_style))
        elements.append(Spacer(1, 6))

        any_meal = False
        for meal_type in (models.MealType.BREAKFAST, models.MealType.LUNCH, models.MealType.DINNER):
            entries = by_date[day][meal_type]
            if not entries:
                continue
            any_meal = True
            label = meal_type_labels[meal_type]
            elements.append(Paragraph(label, meal_heading_style))

            single_no_sub = len(entries) == 1 and not entries[0].sub_category
            if single_no_sub:
                mp = entries[0]
                recipe_title_style = ParagraphStyle(
                    'RecipeTitle',
                    parent=styles['Normal'],
                    fontSize=11,
                    fontName='Helvetica-Bold',
                    spaceAfter=2,
                )
                elements.append(Paragraph(mp.recipe.name, recipe_title_style))
                elements.append(build_recipe_block(mp))
                elements.append(Spacer(1, 8))
            else:
                for mp in entries:
                    elements.append(build_sub_category_block(mp))
                    elements.append(Spacer(1, 6))

        if not any_meal:
            elements.append(Paragraph("<i>Keine Rezepte geplant.</i>", styles['Normal']))

    doc.build(elements)
    filename = f"tageslisten_{sanitize_filename(camp.name)}_{make_timestamp()}.pdf"
    return build_and_serve_pdf(buffer, filename)


@router.get("/recipes/pdf")
async def export_all_recipes_pdf(
    db: Session = Depends(get_db)
):
    """Export all recipes as PDF with improved formatting"""
    recipes = crud.get_recipes(db, skip=0, limit=10000)

    if not recipes:
        raise HTTPException(status_code=404, detail="No recipes found")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles, _, _, section_heading_style = get_pdf_styles()

    title_style = ParagraphStyle(
        'BookTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'BookSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=20,
        alignment=TA_CENTER
    )

    recipe_title_style = ParagraphStyle(
        'RecipeTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=15,
        fontName='Helvetica-Bold'
    )

    # Title page
    elements.append(Paragraph("Rezeptsammlung", title_style))
    elements.append(Paragraph(
        f"{len(recipes)} Rezepte | {datetime.now(timezone.utc).strftime('%d.%m.%Y')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 30))

    # Table of contents
    elements.append(Paragraph("<b>Inhaltsverzeichnis</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))

    toc_data = [["Nr.", "Rezeptname", "Portionen"]]
    for idx, recipe in enumerate(recipes, 1):
        toc_data.append([str(idx), recipe.name, str(recipe.base_servings)])

    toc_table = Table(toc_data, colWidths=[1.5*cm, 11*cm, 3*cm])
    toc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))

    elements.append(toc_table)
    elements.append(PageBreak())

    # Each recipe
    for idx, recipe in enumerate(recipes, 1):
        elements.append(Paragraph(f"{idx}. {recipe.name}", recipe_title_style))

        if recipe.description:
            elements.append(Paragraph(f"<i>{recipe.description}</i>", styles['Normal']))
            elements.append(Spacer(1, 10))

        # Recipe info box
        info_parts = [f"<b>Portionen:</b> {recipe.base_servings}"]
        if recipe.preparation_time:
            info_parts.append(f"<b>Vorbereitung:</b> {recipe.preparation_time} min")
        if recipe.cooking_time:
            info_parts.append(f"<b>Kochzeit:</b> {recipe.cooking_time} min")

        info_table = Table([info_parts], colWidths=[5*cm] * len(info_parts))
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EEF2FF')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 15))

        # Tags and allergens
        if recipe.tags or recipe.allergens:
            parts = []
            if recipe.tags:
                parts.append(f"<b>Tags:</b> {', '.join(t.name for t in recipe.tags)}")
            if recipe.allergens:
                parts.append(f"<b>Allergene:</b> {', '.join(a.name for a in recipe.allergens)}")
            elements.append(Paragraph(" | ".join(parts), styles['Normal']))
            elements.append(Spacer(1, 15))

        # Ingredients
        elements.append(Paragraph("Zutaten", section_heading_style))

        if recipe.ingredients:
            ingredient_data = [["Zutat", "Menge", "Einheit"]]
            for ri in recipe.ingredients:
                ingredient_data.append([ri.ingredient.name, f"{ri.quantity:.1f}", ri.unit])

            ing_table = Table(ingredient_data, colWidths=[10*cm, 3*cm, 3*cm])
            ing_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            elements.append(ing_table)
            elements.append(Spacer(1, 15))

        # Instructions
        if recipe.instructions:
            elements.append(Paragraph("Zubereitung", section_heading_style))
            elements.append(Paragraph(recipe.instructions.replace('\n', '<br/>'), styles['Normal']))

        if idx < len(recipes):
            elements.append(PageBreak())

    doc.build(elements)

    filename = f"rezeptsammlung_{make_timestamp()}.pdf"
    return build_and_serve_pdf(buffer, filename)
